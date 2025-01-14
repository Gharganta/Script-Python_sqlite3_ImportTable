import sqlite3
from openpyxl import load_workbook
conn = sqlite3.connect('utenti.db')
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS utenti (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT NOT NULL,
    cognome TEXT NOT NULL,
    email TEXT NOT NULL,
    numero_telefono TEXT NOT NULL
)
""")

workbook = load_workbook("dati_utenti.xlsx")

foglio = workbook["Utenti"]

for row in foglio.iter_rows(min_row=2, values_only=True):
    nome, cognome, email, numero_telefono = row  
    
    cursor.execute("""
    INSERT INTO utenti (nome, cognome, email, numero_telefono)
    VALUES (?, ?, ?, ?)
    """, (nome, cognome, email, numero_telefono))

conn.commit()

conn.close()

print("Dati inseriti correttamente nel database SQLite 'utenti.db'!")
